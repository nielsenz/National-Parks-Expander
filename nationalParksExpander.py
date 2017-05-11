import os
import openpyxl
import pandas as pd
#Get in excel file
wb = openpyxl.load_workbook('NPS.xlsx')
sheet = wb.get_sheet_by_name('Rec')
coolindex = 2

#Initializing a dataframe in pandas with junk data. 
df = pd.DataFrame({"Park Name" : ["Not Real"],"Park": ["AZJW"], "Date" : [1904], "Attendance" : [394930]}, index = [1])

#Two loop. One to move down the park section
for parks in range(2,376):
	parkCode = 'B' + str(parks)
	parkNameCode = 'A' + str(parks)
	currentPark = sheet[parkCode].value
	currentParkName = sheet[parkNameCode].value
	#And then this should get all the years per park
	for years in range(3,115):
		year = sheet.cell(row=1, column=years).value
		attendance = sheet.cell(row=parks, column=years).value
		if attendance == None:
			attendance = 0
		df2 = pd.DataFrame({"Park Name" : [currentParkName],"Park": [currentPark], "Date" : [year], "Attendance" : [attendance]}, index = [coolindex])
		frames = [df, df2]
		df = pd.concat(frames)
		coolindex=coolindex+1
#drops the first inital row of junk data. 
df.drop(df.head(1).index, inplace=True)
print(df.iloc[:3, :4])
writer = pd.ExcelWriter('output.xlsx')
df.to_excel(writer,sheet_name ='Formatted Data')
writer.save()